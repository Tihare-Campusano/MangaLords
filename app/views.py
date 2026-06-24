from django.shortcuts import render, redirect, get_object_or_404
from .models import Manga, RegistroUsuario, Pedido, DetallePedido, Contacto, Notificacion
from .registroCli import registroClient
from django.contrib.auth import authenticate, login, logout
from .forms import CrudForm, ContactoForm, PerfilForm
from django.contrib.auth.models import User
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.core.mail import send_mail
from django.conf import settings
from django.db import transaction
from django.db.models import Sum
from django.contrib import messages
from django.core.exceptions import ValidationError
import random
import time
import io



# Create your views here.
@login_required
def MangaLords(request):
    if request.user.is_staff:
        return redirect('admin')
    return render(request,'app/MangaLords.html')

@login_required
def admin(request):
    if not request.user.is_staff:
        return redirect('MangaLords')
    
    total_mangas = Manga.objects.count()
    total_stock = Manga.objects.aggregate(Sum('cantidad'))['cantidad__sum'] or 0
    total_users = User.objects.count()
    
    # Obtener algunos mangas recientes para mostrar en el dashboard
    mangas_recientes = Manga.objects.all()[:5]
    
    context = {
        'total_mangas': total_mangas,
        'total_stock': total_stock,
        'total_users': total_users,
        'mangas_recientes': mangas_recientes,
    }
    return render(request, 'app/admin_dashboard.html', context)



def directorio(request):
    if request.user.is_authenticated and request.user.is_staff:
        return redirect('admin')
    from django.db.models import Q
    query = request.GET.get('q', '').strip()
    if query:
        mangas = Manga.objects.filter(
            Q(titulo__icontains=query) | Q(editorial__icontains=query)
        )
    else:
        mangas = Manga.objects.all()
    return render(request,'app/directorio.html',{'mangas': mangas, 'query': query})

def inicioSecion(request):
    form = AuthenticationForm()
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            
            print('Esta correcto')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('MangaLords')
        else:
            print('Incorrecto')
            
    return render(request, 'app/inicioSecion.html', {'login': form})

@login_required
def pagar(request):
    if request.user.is_staff:
        messages.error(request, "Los administradores no pueden realizar compras.")
        return redirect('admin')
    cart = request.session.get('cart', {})
    if not cart:
        messages.warning(request, "Tu carrito está vacío.")
        return redirect('directorio')
        
    # Get profile for pre-filling the shipping/billing forms
    try:
        profile = RegistroUsuario.objects.get(user=request.user)
    except RegistroUsuario.DoesNotExist:
        profile = None

    # Load mangas and validate stock
    mangas_in_cart = []
    total_price = 0
    total_items = 0
    stock_adjusted = False

    for manga_id, quantity in list(cart.items()):
        try:
            manga = Manga.objects.get(pk=manga_id)
            if manga.cantidad == 0:
                del cart[manga_id]
                stock_adjusted = True
                continue
            elif quantity > manga.cantidad:
                cart[manga_id] = manga.cantidad
                quantity = manga.cantidad
                stock_adjusted = True
            
            subtotal = manga.precio * quantity
            total_price += subtotal
            total_items += quantity
            mangas_in_cart.append({
                'manga': manga,
                'quantity': quantity,
                'subtotal': subtotal
            })
        except Manga.DoesNotExist:
            del cart[manga_id]
            stock_adjusted = True

    if stock_adjusted:
        request.session['cart'] = cart
        request.session.modified = True
        messages.warning(request, "El stock de algunos productos en tu carrito ha cambiado. Hemos ajustado las cantidades.")
        if not cart:
            return redirect('directorio')

    if request.method == 'POST':
        # Checkout logic
        nombre = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        telefono = request.POST.get('tel', '').strip()
        direccion = request.POST.get('direccion', '').strip()
        ciudad = request.POST.get('city', '').strip()
        region = request.POST.get('region', '').strip()
        codigo_postal = request.POST.get('cod', '').strip()
        
        # Payment details (just validating they are filled, we don't save card info for security!)
        cardholder = request.POST.get('user', '').strip()
        card_num = request.POST.get('numTar', '').strip()
        
        if not all([nombre, email, telefono, direccion, ciudad, region, cardholder, card_num]):
            messages.error(request, "Por favor rellena todos los campos requeridos.")
        elif not telefono.isdigit() or len(telefono) != 9:
            messages.error(request, "El número de teléfono debe tener exactamente 9 dígitos numéricos.")
        else:
            try:
                card_last_four = card_num[-4:] if len(card_num) >= 4 else card_num
                tx_id = f"ML-{random.randint(100000, 999999)}"
                
                with transaction.atomic():
                    # Create database order first
                    pedido = Pedido.objects.create(
                        user=request.user,
                        transaction_id=tx_id,
                        nombre=nombre,
                        email=email,
                        telefono=telefono,
                        direccion=f"{direccion}, {ciudad}, {region}",
                        total_price=total_price,
                        tarjeta_ultimos_cuatro=card_last_four
                    )
                    
                    # Check and update stock under lock, and create order items
                    items_purchased = []
                    for item in mangas_in_cart:
                        manga = Manga.objects.select_for_update().get(pk=item['manga'].pk)
                        if manga.cantidad < item['quantity']:
                            raise ValidationError(f"Stock insuficiente para {manga.titulo}.")
                        manga.cantidad -= item['quantity']
                        manga.save()
                        
                        # Create detail record
                        DetallePedido.objects.create(
                            pedido=pedido,
                            manga_titulo=manga.titulo,
                            manga_editorial=manga.editorial,
                            precio_unitario=manga.precio,
                            cantidad=item['quantity'],
                            subtotal=item['subtotal'],
                            manga=manga
                        )
                        
                        items_purchased.append({
                            'titulo': manga.titulo,
                            'editorial': manga.editorial,
                            'precio': manga.precio,
                            'quantity': item['quantity'],
                            'subtotal': item['subtotal']
                        })
                    
                    # Store purchase details in session for the receipt page
                    request.session['receipt'] = {
                        'items': items_purchased,
                        'total_price': total_price,
                        'nombre': nombre,
                        'direccion': f"{direccion}, {ciudad}, {region}",
                        'transaction_id': tx_id
                    }
                    
                    # Clear cart
                    request.session['cart'] = {}
                    request.session.modified = True
                    
                    # Notificar al cliente
                    Notificacion.objects.create(
                        user=request.user,
                        mensaje="🔔 Tu compra fue realizada exitosamente."
                    )
                    
                    return redirect('pago_exitoso')
            except Exception as e:
                messages.error(request, f"Ocurrió un error al procesar tu pago: {str(e)}")

    context = {
        'mangas': mangas_in_cart,
        'total_price': total_price,
        'total_items': total_items,
        'profile': profile,
    }
    return render(request, 'app/pagar.html', context)

@login_required
def cerrarSesion(request):
    logout(request)
    return redirect('inicioSecion')

# aca se usa esto iregistroSecion
def Registro(request):
    form = registroClient()
    if request.method == 'POST':
        form = registroClient(data=request.POST)
        if form.is_valid():
            email = form.cleaned_data.get('email')
            password1 = form.cleaned_data.get('password1')
            nombre = form.cleaned_data.get('nombres')
            apellido = form.cleaned_data.get('apellidos')
            telefono = form.cleaned_data.get('telefono')
            
            # Verificar de forma segura si el email o el usuario ya existe
            if User.objects.filter(username=email).exists() or User.objects.filter(email=email).exists():
                form.add_error('email', 'Este correo electrónico ya está registrado.')
            else:
                user = User.objects.create_user(
                    username=email,
                    password=password1,
                    email=email,
                    first_name=nombre,
                    last_name=apellido
                )
                user.save()
                usuario = RegistroUsuario.objects.create(
                    user=user,
                    nombres=nombre,
                    apellidos=apellido,
                    email=email,
                    telefono=telefono
                )
                usuario.save()
                
                # Notificar a los administradores
                for admin_user in User.objects.filter(is_staff=True):
                    Notificacion.objects.create(
                        user=admin_user,
                        mensaje=f"🔔 Nuevo usuario registrado: {nombre} {apellido}."
                    )

                print('Se guardo')
                login(request, user=user)
                return redirect('MangaLords')
        else:
            print(form.errors)
            
    return render(request, 'app/RegistroUsuario.html', {'registroCli': form})


# contacto
@login_required
def contacto(request):
    if request.user.is_staff:
        mensajes = Contacto.objects.all().order_by('-id')
        return render(request, 'app/contacto_admin.html', {'mensajes': mensajes})

    initial_data = {}
    registro = RegistroUsuario.objects.filter(user=request.user).first()
    if registro:
        initial_data['nombre'] = f"{registro.nombres} {registro.apellidos}"[:50]
        initial_data['email'] = registro.email
        initial_data['telefono'] = registro.telefono
    else:
        initial_data['nombre'] = request.user.get_full_name() or request.user.username
        initial_data['email'] = request.user.email

    from django.db.models import Q
    user_consultas = Contacto.objects.filter(
        Q(user=request.user) | Q(email=request.user.email) | Q(email=request.user.username)
    ).order_by('-fecha_creacion')

    if request.method == 'POST':
        formulario = ContactoForm(data=request.POST, initial=initial_data)
        formulario.fields['nombre'].disabled = True
        formulario.fields['email'].disabled = True
        
        if formulario.is_valid():
            contacto_obj = formulario.save(commit=False)
            contacto_obj.user = request.user
            contacto_obj.save()
            
            # Notificar a los administradores
            for admin_user in User.objects.filter(is_staff=True):
                Notificacion.objects.create(
                    user=admin_user,
                    mensaje=f"🔔 Nuevo mensaje pendiente de respuesta de {contacto_obj.nombre}."
                )
            
            form_nuevo = ContactoForm(initial=initial_data)
            form_nuevo.fields['nombre'].disabled = True
            form_nuevo.fields['email'].disabled = True
            
            user_consultas_new = Contacto.objects.filter(
                Q(user=request.user) | Q(email=request.user.email) | Q(email=request.user.username)
            ).order_by('-fecha_creacion')
            
            data = {
                'form': form_nuevo,
                'mensaje': "Mensaje enviado",
                'consultas': user_consultas_new
            }
            return render(request, 'app/contacto.html', data)
        else:
            data = {
                'form': formulario,
                'consultas': user_consultas
            }
    else:
        formulario = ContactoForm(initial=initial_data)
        formulario.fields['nombre'].disabled = True
        formulario.fields['email'].disabled = True
        data = {
            'form': formulario,
            'consultas': user_consultas
        }
            
    return render(request,'app/contacto.html',data)


# leer datos de la clase creada
def vistaManga(request,pk):
    mangas = Manga.objects.get(id=pk)
    return render(request,'app/vistaManga.html',{'mangas': mangas})
    
    

# hecho por el make agregar listar modificar y eliminar
@login_required
def agregar_producto(request):
    if not request.user.is_staff:
        return redirect('MangaLords')
    
    data = {
        'form': CrudForm()
    }

    if request.method == 'POST':
        # Verificar si es una importacion de Excel
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                data['error_excel'] = 'Por favor, sube un archivo Excel válido (.xlsx o .xls)'
            else:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(excel_file)
                    sheet = wb.active
                    
                    mangas_creados = 0
                    mangas_actualizados = 0
                    errores = []
                    
                    with transaction.atomic():
                        # Comenzamos desde la fila 2 (la fila 1 tiene cabeceras)
                        for r_idx in range(2, sheet.max_row + 1):
                            # Columnas: 1: ID, 2: Título, 3: Editorial, 4: Precio, 5: Descripción, 6: Cantidad
                            m_id = sheet.cell(row=r_idx, column=1).value
                            m_titulo = sheet.cell(row=r_idx, column=2).value
                            m_editorial = sheet.cell(row=r_idx, column=3).value
                            m_precio = sheet.cell(row=r_idx, column=4).value
                            m_desc = sheet.cell(row=r_idx, column=5).value
                            m_cant = sheet.cell(row=r_idx, column=6).value
                            
                            # Saltar filas completamente vacías
                            if m_id is None and m_titulo is None:
                                continue
                            
                            # Formatear el ID de forma segura (sin decimales .0)
                            if m_id is not None:
                                try:
                                    m_id_str = str(int(float(str(m_id)))).strip()
                                except ValueError:
                                    m_id_str = str(m_id).strip()
                            else:
                                m_id_str = ""
                                
                            # Validaciones básicas
                            if not m_id_str or len(m_id_str) > 6:
                                errores.append(f"Fila {r_idx}: ID '{m_id_str}' no válido o supera los 6 caracteres.")
                                continue
                            if not m_titulo:
                                errores.append(f"Fila {r_idx}: El título es obligatorio.")
                                continue
                                
                            # Parsear valores numéricos de forma segura
                            try:
                                precio_parsed = int(float(str(m_precio).strip())) if m_precio is not None else 0
                                if precio_parsed < 0:
                                    precio_parsed = 0
                                    errores.append(f"Fila {r_idx}: El precio no puede ser negativo, se ajustó a $0.")
                            except ValueError:
                                precio_parsed = 0
                                errores.append(f"Fila {r_idx}: Precio '{m_precio}' no es un número válido, se ajustó a $0.")
                                
                            try:
                                cant_parsed = int(float(str(m_cant).strip())) if m_cant is not None else 0
                                if cant_parsed < 0:
                                    cant_parsed = 0
                                    errores.append(f"Fila {r_idx}: El stock no puede ser negativo, se ajustó a 0.")
                            except ValueError:
                                cant_parsed = 0
                                errores.append(f"Fila {r_idx}: Cantidad '{m_cant}' no es un número válido, se ajustó a 0.")
                                
                            manga_obj, created = Manga.objects.update_or_create(
                                id=m_id_str,
                                defaults={
                                    'titulo': str(m_titulo).strip()[:50],
                                    'editorial': str(m_editorial).strip()[:50] if m_editorial else '',
                                    'precio': precio_parsed,
                                    'descripcion': str(m_desc).strip() if m_desc else '',
                                    'cantidad': cant_parsed
                                }
                            )
                            
                            if created:
                                mangas_creados += 1
                            else:
                                mangas_actualizados += 1
                                
                    if mangas_creados > 0:
                        for u in User.objects.filter(is_staff=False):
                            Notificacion.objects.create(
                                user=u,
                                mensaje="🔔 Se agregaron nuevos mangas."
                            )
                                
                    if errores:
                        data['error_excel'] = "Errores en la carga: " + " | ".join(errores[:3])
                        if len(errores) > 3:
                            data['error_excel'] += f" (...y {len(errores) - 3} errores más)"
                    data['mensaje_excel'] = f"Importación masiva completada. Creados: {mangas_creados}, Actualizados: {mangas_actualizados}."
                    
                except Exception as e:
                    data['error_excel'] = f"Error general al procesar el Excel: {str(e)}"
        else:
            # Formulario manual normal
            formulario = CrudForm(request.POST, request.FILES)
            if formulario.is_valid():
                manga = formulario.save()
                data["mensaje"] = "Manga guardado correctamente"
                
                # Notificar a los clientes
                for u in User.objects.filter(is_staff=False):
                    Notificacion.objects.create(
                        user=u,
                        mensaje=f"🔔 Se agregaron nuevos mangas: {manga.titulo}."
                    )
                    if manga.descuento > 0:
                        Notificacion.objects.create(
                            user=u,
                            mensaje=f"🔔 Oferta disponible: {manga.descuento}% de descuento en {manga.titulo}."
                        )
            else:
                data["form"] = formulario
                print(formulario.errors)

    return render(request, 'app/crud/agregar.html', data)

@login_required
def listar_productos(request):
    if not request.user.is_staff:
        return redirect('MangaLords')
        
    mangas = Manga.objects.all()
    
    # Obtener todos los usuarios de forma segura
    usuarios = []
    for u in User.objects.all():
        profile = getattr(u, 'registrousuario', None)
        if profile:
            nombres = f"{profile.nombres} {profile.apellidos}"
            email = profile.email
            telefono = profile.telefono
        else:
            nombres = f"{u.first_name} {u.last_name}".strip() or u.username
            email = u.email
            telefono = '-'
            
        usuarios.append({
            'username': u.username,
            'nombres': nombres,
            'email': email,
            'telefono': telefono,
            'is_staff': u.is_staff,
        })

    data = {
        'mangas': mangas,
        'usuarios': usuarios
    }

    return render(request, 'app/crud/listar.html', data)

@login_required
def modificar_manga(request, id):
    if not request.user.is_staff:
        return redirect('MangaLords')

    manga = get_object_or_404(Manga, id=id)

    data = {
        'form': CrudForm(instance=manga),
        'manga': manga
    }

    if request.method == 'POST':
        formulario = CrudForm(data=request.POST, instance=manga, files=request.FILES)

        if formulario.is_valid():
            original_descuento = manga.descuento
            manga_saved = formulario.save()
            
            if manga_saved.descuento > original_descuento and manga_saved.descuento > 0:
                for u in User.objects.filter(is_staff=False):
                    Notificacion.objects.create(
                        user=u,
                        mensaje=f"🔔 Oferta disponible: {manga_saved.descuento}% de descuento en {manga_saved.titulo}."
                    )
            return redirect(to="listar_productos")
        data["form"] = formulario

    return render(request, 'app/crud/modificar.html', data)

@login_required
def eliminar_producto(request, id):
    if not request.user.is_staff:
        return redirect('MangaLords')
        
    manga = get_object_or_404(Manga, id=id)
    manga.delete()
    return redirect(to="listar_productos")



def carrito(request):
    if request.user.is_authenticated and request.user.is_staff:
        messages.warning(request, "Los administradores no tienen acceso al carrito de compras.")
        return redirect('admin')
    cart = request.session.get('cart', {})
    mangas_in_cart = []
    total_price = 0
    total_items = 0
    stock_adjusted = False

    for manga_id, quantity in list(cart.items()):
        try:
            manga = Manga.objects.get(pk=manga_id)
            if manga.cantidad == 0:
                del cart[manga_id]
                stock_adjusted = True
                continue
            elif quantity > manga.cantidad:
                cart[manga_id] = manga.cantidad
                quantity = manga.cantidad
                stock_adjusted = True
                
            subtotal = manga.precio * quantity
            total_price += subtotal
            total_items += quantity
            mangas_in_cart.append({
                'manga': manga,
                'quantity': quantity,
                'subtotal': subtotal
            })
        except Manga.DoesNotExist:
            del cart[manga_id]
            stock_adjusted = True

    if stock_adjusted:
        request.session['cart'] = cart
        request.session.modified = True
        messages.warning(request, "El inventario de algunos mangas cambió. Hemos ajustado tu carrito.")

    context = {
        'mangas': mangas_in_cart,
        'total_price': total_price,
        'total_items': total_items,
    }
    return render(request, 'app/carrito.html', context)


@login_required
def perfil(request):
    registro, created = RegistroUsuario.objects.get_or_create(
        user=request.user,
        defaults={
            'nombres': request.user.first_name or request.user.username,
            'apellidos': request.user.last_name or '',
            'email': request.user.email or '',
            'telefono': 999999999
        }
    )
    
    mensaje = None
    
    if request.method == 'POST':
        form = PerfilForm(request.POST, instance=registro, user=request.user)
        if form.is_valid():
            registro = form.save()
            
            user = request.user
            user.first_name = registro.nombres
            user.last_name = registro.apellidos
            user.email = registro.email
            user.save()
            
            mensaje = "Tu perfil ha sido actualizado con éxito."
    else:
        form = PerfilForm(instance=registro, user=request.user)
        
    return render(request, 'app/perfil.html', {'form': form, 'mensaje': mensaje})


# ─── PASSWORD RESET OTP ────────────────────────────────────────────────────────

@require_POST
def request_password_reset(request):
    """Genera un código OTP de 6 dígitos y lo guarda en sesión."""
    email = request.POST.get('email', '').strip().lower()
    
    if not email:
        return JsonResponse({'success': False, 'error': 'Ingresa tu correo electrónico.'})
    
    # Verificar si el correo existe (filter evita MultipleObjectsReturned)
    user = User.objects.filter(email=email).first()
    if user is None:
        # Por seguridad, no revelamos si el correo existe o no
        return JsonResponse({'success': True, 'message': 'Si el correo existe, recibirás un código.'})
    
    # Generar OTP de 6 dígitos
    otp_code = str(random.randint(100000, 999999))
    otp_expiry = time.time() + getattr(settings, 'PASSWORD_RESET_OTP_EXPIRY', 600)
    
    # Guardar en sesión
    request.session['pwd_reset_otp'] = otp_code
    request.session['pwd_reset_otp_expiry'] = otp_expiry
    request.session['pwd_reset_email'] = email
    request.session.modified = True
    
    # Enviar correo (en desarrollo aparece en consola Y en el modal)
    try:
        send_mail(
            subject='🔐 MangaLords – Tu código de verificación',
            message=f'Tu código de verificación es: {otp_code}\n\nEste código expira en 10 minutos.\n\nSi no solicitaste esto, ignora este mensaje.',
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            fail_silently=True,
        )
    except Exception:
        pass

    response_data = {'success': True, 'message': 'Si el correo existe, recibirás un código.'}

    # En modo DEBUG mostramos el código directamente (solo desarrollo)
    if settings.DEBUG:
        response_data['debug_code'] = otp_code

    return JsonResponse(response_data)


@require_POST
def verify_reset_code(request):
    """Verifica el código OTP ingresado por el usuario."""
    code = request.POST.get('code', '').strip()
    
    otp_code = request.session.get('pwd_reset_otp')
    otp_expiry = request.session.get('pwd_reset_otp_expiry', 0)
    
    if not otp_code:
        return JsonResponse({'success': False, 'error': 'No hay código pendiente. Solicita uno nuevo.'})
    
    if time.time() > otp_expiry:
        # Limpiar sesión expirada
        for key in ['pwd_reset_otp', 'pwd_reset_otp_expiry', 'pwd_reset_email']:
            request.session.pop(key, None)
        return JsonResponse({'success': False, 'error': 'El código ha expirado. Solicita uno nuevo.'})
    
    if code != otp_code:
        return JsonResponse({'success': False, 'error': 'Código incorrecto. Inténtalo de nuevo.'})
    
    # Marcar como verificado
    request.session['pwd_reset_verified'] = True
    request.session.modified = True
    return JsonResponse({'success': True})


@require_POST
def reset_password(request):
    """Actualiza la contraseña del usuario tras verificación OTP."""
    if not request.session.get('pwd_reset_verified'):
        return JsonResponse({'success': False, 'error': 'No autorizado. Verifica tu código primero.'})
    
    email = request.session.get('pwd_reset_email')
    password1 = request.POST.get('password1', '')
    password2 = request.POST.get('password2', '')
    
    if not password1 or len(password1) < 8:
        return JsonResponse({'success': False, 'error': 'La contraseña debe tener al menos 8 caracteres.'})
    
    if password1 != password2:
        return JsonResponse({'success': False, 'error': 'Las contraseñas no coinciden.'})
    
    user = User.objects.filter(email=email).first()
    if user is None:
        return JsonResponse({'success': False, 'error': 'Usuario no encontrado.'})
    
    user.set_password(password1)
    user.save()
    
    # Limpiar sesión de reset
    for key in ['pwd_reset_otp', 'pwd_reset_otp_expiry', 'pwd_reset_email', 'pwd_reset_verified']:
        request.session.pop(key, None)
    
    return JsonResponse({'success': True, 'message': '¡Contraseña actualizada correctamente!'})


def search_suggestions(request):
    from django.http import JsonResponse
    from django.db.models import Q
    
    query = request.GET.get('q', '').strip()
    if len(query) >= 1:
        mangas = Manga.objects.filter(
            Q(titulo__icontains=query) | Q(editorial__icontains=query)
        )[:6]
        results = []
        for m in mangas:
            results.append({
                'id': m.id,
                'titulo': m.titulo,
                'editorial': m.editorial,
                'imagen_url': m.imagen.url if m.imagen else '',
            })
        return JsonResponse({'success': True, 'results': results})
    return JsonResponse({'success': True, 'results': []})


# ─── EXPORTACIONES E IMPORTACIONES (EXCEL Y PDF) ────────────────────────────────

from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors

class NumberedCanvas(canvas.Canvas):
    """Lienzo personalizado para dibujar cabecera, pie de página y número de páginas dinámico."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_elements(num_pages)
            super().showPage()
        super().save()

    def draw_page_elements(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#718096"))
        
        # Dibujar cabecera
        self.setLineWidth(0.5)
        self.setStrokeColor(colors.HexColor("#CBD5E0"))
        self.line(36, 756, 576, 756)
        self.drawString(36, 762, "MangaLords – Panel de Administración")
        
        # Dibujar pie de página
        self.line(36, 45, 576, 45)
        self.drawString(36, 32, "© 2026 MangaLords. Todos los derechos reservados.")
        
        page_text = f"Página {self._pageNumber} de {page_count}"
        self.drawRightString(576, 32, page_text)
        self.restoreState()


@login_required
def download_manga_template(request):
    """Genera una plantilla de Excel (.xlsx) con el formato requerido para importar mangas."""
    if not request.user.is_staff:
        return redirect('MangaLords')
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Mangas"
    
    headers = ["ID", "Título", "Editorial", "Precio", "Descripción", "Cantidad"]
    ws.append(headers)
    
    # Estilos de cabecera
    header_fill = PatternFill(start_color="185E83", end_color="185E83", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num in range(1, 7):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        
    # Fila de ejemplo
    ws.append(["M00001", "Dragon Ball Vol. 1", "Panini", 8990, "Primera edición del clásico manga shonen.", 15])
    
    # Ajuste de tamaño de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="plantilla_mangas.xlsx"'
    wb.save(response)
    return response


@login_required
def export_mangas_excel(request):
    """Exporta el listado completo de mangas a un archivo Excel (.xlsx) con formato."""
    if not request.user.is_staff:
        return redirect('MangaLords')
        
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Mangas"
    
    # Fila de título fusionada
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "MangaLords – Reporte Completo de Inventario"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="121216", end_color="121216", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    ws.append([]) # Fila vacía de separación
    
    # Cabeceras de la tabla
    headers = ["ID", "Título", "Editorial", "Precio", "Descripción", "Cantidad (Stock)"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="C91A24", end_color="C91A24", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 24
    
    for col_num in range(1, 7):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Insertar registros de mangas
    for manga in Manga.objects.all():
        ws.append([
            manga.id,
            manga.titulo,
            manga.editorial,
            manga.precio,
            manga.descripcion,
            manga.cantidad
        ])
        
        r_idx = ws.max_row
        # Formatos numéricos
        ws.cell(row=r_idx, column=4).number_format = '$#,##0'
        ws.cell(row=r_idx, column=6).number_format = '#,##0'
        
        for col_num in range(1, 7):
            cell = ws.cell(row=r_idx, column=col_num)
            cell.border = thin_border
            cell.font = Font(name="Arial", size=10)
            if col_num in [1, 4, 6]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")
                
    # Autoajuste de columnas
    for col in ws.columns:
        cells_to_check = [cell.value for cell in col if cell.row > 1]
        max_len = max(len(str(v or '')) for v in cells_to_check) if cells_to_check else 10
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40) # limitar a ancho máximo 40
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="reporte_mangas.xlsx"'
    wb.save(response)
    return response


@login_required
def export_users_excel(request):
    """Exporta el listado de usuarios registrados (sin contraseñas) a un archivo Excel (.xlsx)."""
    if not request.user.is_staff:
        return redirect('MangaLords')
        
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista Usuarios"
    
    # Fila de título
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = "MangaLords – Listado de Usuarios Registrados"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="121216", end_color="121216", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    ws.append([]) # Separador
    
    headers = ["Username/Email", "Nombre Completo", "Correo Electrónico", "Teléfono"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="185E83", end_color="185E83", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 24
    
    for col_num in range(1, 5):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    for u in User.objects.all():
        profile = getattr(u, 'registrousuario', None)
        if profile:
            nombres = f"{profile.nombres} {profile.apellidos}"
            email = profile.email
            telefono = profile.telefono
        else:
            nombres = f"{u.first_name} {u.last_name}".strip() or u.username
            email = u.email
            telefono = '-'
            
        ws.append([u.username, nombres, email, telefono])
        r_idx = ws.max_row
        for col_num in range(1, 5):
            cell = ws.cell(row=r_idx, column=col_num)
            cell.border = thin_border
            cell.font = Font(name="Arial", size=10)
            if col_num in [1, 4]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")
                
    for col in ws.columns:
        cells_to_check = [cell.value for cell in col if cell.row > 1]
        max_len = max(len(str(v or '')) for v in cells_to_check) if cells_to_check else 10
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 15), 35)
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="reporte_usuarios.xlsx"'
    wb.save(response)
    return response


@login_required
def export_mangas_pdf(request):
    """Genera un archivo PDF con la lista completa de mangas formateada con ReportLab."""
    if not request.user.is_staff:
        return redirect('MangaLords')
        
    import io
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=54,
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        name='ReportTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor("#C91A24"),
        spaceAfter=15
    )
    
    cell_header_style = ParagraphStyle(
        name='CellHeader',
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=1
    )
    
    cell_text_style = ParagraphStyle(
        name='CellText',
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#2D3748")
    )
    
    cell_center_style = ParagraphStyle(
        name='CellCenter',
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#2D3748"),
        alignment=1
    )
    
    story = []
    story.append(Spacer(1, 15))
    story.append(Paragraph("Inventario Completo de Mangas", title_style))
    story.append(Spacer(1, 10))
    
    headers = [
        Paragraph("ID", cell_header_style),
        Paragraph("Título del Manga", cell_header_style),
        Paragraph("Editorial", cell_header_style),
        Paragraph("Precio", cell_header_style),
        Paragraph("Descripción", cell_header_style),
        Paragraph("Stock", cell_header_style)
    ]
    
    table_data = [headers]
    
    for manga in Manga.objects.all():
        desc_corta = manga.descripcion
        if len(desc_corta) > 120:
            desc_corta = desc_corta[:120] + "..."
            
        table_data.append([
            Paragraph(manga.id, cell_center_style),
            Paragraph(manga.titulo, cell_text_style),
            Paragraph(manga.editorial, cell_text_style),
            Paragraph(f"${manga.precio:,}", cell_center_style),
            Paragraph(desc_corta or "-", cell_text_style),
            Paragraph(str(manga.cantidad), cell_center_style)
        ])
        
    col_widths = [50, 110, 80, 60, 190, 50] # Suma = 540 (ancho total de página letter horizontalmente libre)
    
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#C91A24")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor("#FFFFFF"), colors.HexColor("#F7FAFC")]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ('TOPPADDING', (0,1), (-1,-1), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
    ]))
    
    story.append(t)
    doc.build(story, canvasmaker=NumberedCanvas)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="inventario_mangas.pdf"'
    response.write(buffer.getvalue())
    buffer.close()
    return response


@login_required
def export_users_pdf(request):
    """Genera un archivo PDF con el listado seguro de usuarios registrados."""
    if not request.user.is_staff:
        return redirect('MangaLords')
        
    import io
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=54,
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        name='ReportTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor("#185E83"),
        spaceAfter=15
    )
    
    cell_header_style = ParagraphStyle(
        name='CellHeader',
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.white,
        alignment=1
    )
    
    cell_text_style = ParagraphStyle(
        name='CellText',
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor("#2D3748")
    )
    
    cell_center_style = ParagraphStyle(
        name='CellCenter',
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor("#2D3748"),
        alignment=1
    )
    
    story = []
    story.append(Spacer(1, 15))
    story.append(Paragraph("Usuarios Registrados en MangaLords", title_style))
    story.append(Spacer(1, 10))
    
    headers = [
        Paragraph("Nombre Completo", cell_header_style),
        Paragraph("Correo Electrónico", cell_header_style),
        Paragraph("Teléfono de Contacto", cell_header_style)
    ]
    
    table_data = [headers]
    
    for u in User.objects.all():
        profile = getattr(u, 'registrousuario', None)
        if profile:
            nombres = f"{profile.nombres} {profile.apellidos}"
            email = profile.email
            telefono = str(profile.telefono)
        else:
            nombres = f"{u.first_name} {u.last_name}".strip() or u.username
            email = u.email
            telefono = '-'
            
        table_data.append([
            Paragraph(nombres, cell_text_style),
            Paragraph(email, cell_text_style),
            Paragraph(telefono, cell_center_style)
        ])
        
    col_widths = [200, 220, 120]
    
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#185E83")),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor("#FFFFFF"), colors.HexColor("#F7FAFC")]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ('TOPPADDING', (0,1), (-1,-1), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
    ]))
    
    story.append(t)
    doc.build(story, canvasmaker=NumberedCanvas)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_usuarios.pdf"'
    response.write(buffer.getvalue())
    buffer.close()
    return response


@login_required
def agregar_al_carrito(request, pk):
    if request.user.is_staff:
        messages.error(request, "Los administradores no pueden añadir productos al carrito.")
        return redirect('admin')
    manga = get_object_or_404(Manga, pk=pk)
    
    # Check if stock is 0
    if manga.cantidad <= 0:
        messages.error(request, f"Lo sentimos, {manga.titulo} está temporalmente sin stock.")
        return redirect('directorio')
        
    cart = request.session.get('cart', {})
    current_qty = cart.get(pk, 0)
    
    if current_qty + 1 > manga.cantidad:
        messages.warning(request, f"No puedes agregar más unidades de {manga.titulo}. Stock máximo disponible: {manga.cantidad}.")
    else:
        cart[pk] = current_qty + 1
        request.session['cart'] = cart
        request.session.modified = True
        messages.success(request, f"¡{manga.titulo} añadido al carrito!")
        
    return redirect('carrito')


@login_required
def eliminar_del_carrito(request, pk):
    if request.user.is_staff:
        messages.error(request, "Acción no permitida para administradores.")
        return redirect('admin')
    cart = request.session.get('cart', {})
    if pk in cart:
        del cart[pk]
        request.session['cart'] = cart
        request.session.modified = True
        messages.success(request, "Producto eliminado del carrito.")
    return redirect('carrito')


@login_required
@require_POST
def actualizar_cantidad_carrito(request, pk):
    if request.user.is_staff:
        messages.error(request, "Acción no permitida para administradores.")
        return redirect('admin')
    manga = get_object_or_404(Manga, pk=pk)
    cart = request.session.get('cart', {})
    
    try:
        quantity = int(request.POST.get('quantity', 1))
    except (ValueError, TypeError):
        quantity = 1
        
    if quantity <= 0:
        if pk in cart:
            del cart[pk]
            messages.success(request, "Producto eliminado del carrito.")
    else:
        if quantity > manga.cantidad:
            cart[pk] = manga.cantidad
            messages.warning(request, f"Cantidad ajustada al stock máximo disponible para {manga.titulo} ({manga.cantidad} unidades).")
        else:
            cart[pk] = quantity
            messages.success(request, f"Cantidad de {manga.titulo} actualizada.")
            
    request.session['cart'] = cart
    request.session.modified = True
    return redirect('carrito')


@login_required
def pago_exitoso(request):
    receipt = request.session.get('receipt', None)
    if not receipt:
        return redirect('directorio')
        
    context = {
        'receipt': receipt
    }
    # Del receipt from session so they cannot reload and see the boleta again with same data
    if 'receipt' in request.session:
        del request.session['receipt']
        request.session.modified = True
        
    return render(request, 'app/pago_exitoso.html', context)


@login_required
def compras_realizadas(request):
    if request.user.is_staff:
        pedidos = Pedido.objects.all().order_by('-fecha').prefetch_related('detalles__manga')
    else:
        pedidos = Pedido.objects.filter(user=request.user).order_by('-fecha').prefetch_related('detalles__manga')
    context = {
        'pedidos': pedidos
    }
    return render(request, 'app/compras_realizadas.html', context)


@login_required
def toggle_user_role(request, username):
    if not request.user.is_staff:
        return redirect('MangaLords')
    
    target_user = get_object_or_404(User, username=username)
    if target_user == request.user:
        messages.error(request, "No puedes cambiar tu propio rol para no perder el acceso a la administración.")
    else:
        target_user.is_staff = not target_user.is_staff
        target_user.save()
        role_str = "Administrador" if target_user.is_staff else "Cliente"
        messages.success(request, f"El rol del usuario {target_user.username} ha sido cambiado a {role_str} correctamente.")
        
    return redirect('listar_productos')


@login_required
def responder_contacto(request, pk):
    if not request.user.is_staff:
        return redirect('MangaLords')
        
    contacto_obj = get_object_or_404(Contacto, pk=pk)
    
    if request.method == 'POST':
        respuesta_text = request.POST.get('respuesta', '').strip()
        if respuesta_text:
            from django.utils import timezone
            contacto_obj.respuesta = respuesta_text
            contacto_obj.respondido = True
            contacto_obj.fecha_respuesta = timezone.now()
            contacto_obj.save()
            
            # Notificar al cliente
            target_user = contacto_obj.user
            if not target_user and contacto_obj.email:
                target_user = User.objects.filter(username=contacto_obj.email).first() or User.objects.filter(email=contacto_obj.email).first()
                
            if target_user:
                Notificacion.objects.create(
                    user=target_user,
                    mensaje="🔔 Tu mensaje fue respondido."
                )

            messages.success(request, f"Respuesta enviada con éxito a {contacto_obj.nombre}.")
        else:
            messages.error(request, "La respuesta no puede estar vacía.")
            
    return redirect('contacto')


@login_required
def get_notifications(request):
    from django.http import JsonResponse
    notifs = Notificacion.objects.filter(user=request.user).order_by('-fecha_creacion')
    data = []
    for n in notifs[:20]:
        data.append({
            'id': n.id,
            'mensaje': n.mensaje,
            'leido': n.leido,
            'fecha_creacion': n.fecha_creacion.strftime('%d/%m/%Y %H:%M')
        })
    unread_count = notifs.filter(leido=False).count()
    return JsonResponse({
        'notifications': data,
        'unread_count': unread_count
    })


@login_required
def mark_notification_read(request, pk):
    from django.http import JsonResponse
    if request.method == 'POST':
        notif = get_object_or_404(Notificacion, pk=pk, user=request.user)
        notif.leido = True
        notif.save()
        unread_count = Notificacion.objects.filter(user=request.user, leido=False).count()
        return JsonResponse({'success': True, 'unread_count': unread_count})
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=400)


@login_required
def mark_all_notifications_read(request):
    from django.http import JsonResponse
    if request.method == 'POST':
        Notificacion.objects.filter(user=request.user, leido=False).update(leido=True)
        return JsonResponse({'success': True, 'unread_count': 0})
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=400)


